from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.core.checks import run_checks
from django.core.checks.registry import registry as django_checks_registry
from django.core.cache import cache
from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from general_manager import bootstrap as gm_bootstrap
from general_manager.cache.cache_decorator import cached
from general_manager.cache.dependency_index import (
    begin_dependency_data_change,
    end_dependency_data_change,
    get_dependency_generation,
    is_dependency_data_change_active,
)
from general_manager.interface import ExcelInterface
from general_manager.interface.excel import (
    ExcelCharField,
    ExcelDecimalField,
    ExcelValidationError,
    ExcelField,
    ExcelIntegerField,
    ExcelMutationPayloadError,
    ExcelStructureError,
    ExcelSyncDelta,
    ExcelWriteConflictError,
)
from general_manager.interface.excel_store import DEFAULT_EXCEL_STORE
from general_manager.interface.capabilities.excel import EXCEL_SYSTEM_CHECK_ID
from general_manager.interface.infrastructure.startup_hooks import (
    clear_startup_hooks,
    registered_startup_hooks,
)
from general_manager.interface.infrastructure.system_checks import (
    clear_system_checks,
    registered_system_checks,
)
from general_manager.interface.interfaces import (
    ExcelCharField as InterfacesExcelCharField,
    ExcelDecimalField as InterfacesExcelDecimalField,
    ExcelField as InterfacesExcelField,
    ExcelIntegerField as InterfacesExcelIntegerField,
)
from general_manager.manager.general_manager import GeneralManager
from general_manager.utils.make_cache_key import make_cache_key


class ExcelFieldTests(SimpleTestCase):
    def test_public_imports_are_available(self) -> None:
        self.assertTrue(issubclass(ExcelInterface, object))
        self.assertIs(ExcelCharField(max_length=12).python_type, str)

    def test_concrete_interface_module_exports_excel_fields(self) -> None:
        self.assertIs(InterfacesExcelField, ExcelField)
        self.assertIs(InterfacesExcelCharField, ExcelCharField)
        self.assertIs(InterfacesExcelIntegerField, ExcelIntegerField)
        self.assertIs(InterfacesExcelDecimalField, ExcelDecimalField)

    def test_field_parse_and_dump(self) -> None:
        char_field = ExcelCharField(max_length=4)
        int_field = ExcelIntegerField()
        decimal_field = ExcelDecimalField(max_digits=6, decimal_places=2)

        self.assertEqual(char_field.parse("AB"), "AB")
        self.assertEqual(int_field.parse("5"), 5)
        self.assertEqual(decimal_field.parse("10.50"), Decimal("10.50"))
        self.assertEqual(decimal_field.dump(Decimal("3.25")), Decimal("3.25"))

    def test_decimal_field_rejects_whole_number_max_digits_overflow(self) -> None:
        field = ExcelDecimalField(max_digits=3)

        with self.assertRaises(ExcelValidationError):
            field.parse("1000")

    def test_decimal_field_checks_max_digits_after_quantization(self) -> None:
        field = ExcelDecimalField(max_digits=4, decimal_places=2)

        with self.assertRaises(ExcelValidationError):
            field.parse("99.995")

    def test_fields_compare_by_identity(self) -> None:
        char_field = ExcelCharField(max_length=4)
        same_char_constraints = ExcelCharField(max_length=4)
        different_char_constraints = ExcelCharField(max_length=5)
        decimal_field = ExcelDecimalField(max_digits=4, decimal_places=2)
        same_decimal_constraints = ExcelDecimalField(max_digits=4, decimal_places=2)

        self.assertNotEqual(char_field, same_char_constraints)
        self.assertNotEqual(char_field, different_char_constraints)
        self.assertNotEqual(decimal_field, same_decimal_constraints)
        self.assertEqual(
            len(
                {
                    char_field,
                    same_char_constraints,
                    different_char_constraints,
                    decimal_field,
                    same_decimal_constraints,
                }
            ),
            5,
        )

    def test_required_field_rejects_blank_value(self) -> None:
        field = ExcelField(str, required=True)

        with self.assertRaises(ExcelValidationError):
            field.parse("")


class ExcelProduct(GeneralManager):
    sku: str
    name: str

    class Interface(ExcelInterface):
        sku = ExcelCharField(max_length=12, unique=True)
        name = ExcelCharField(max_length=80)

        class Meta:
            workbook = "products.xlsx"
            sheet = "Products"
            table = "ProductsTable"
            key = "sku"


class ExcelLifecycleTests(SimpleTestCase):
    def test_meta_and_fields_are_normalized(self) -> None:
        self.assertEqual(ExcelProduct.Interface._interface_type, "excel")
        self.assertEqual(ExcelProduct.Interface.excel_meta.workbook, "products.xlsx")
        self.assertEqual(ExcelProduct.Interface.excel_meta.sheet, "Products")
        self.assertEqual(ExcelProduct.Interface.excel_meta.table, "ProductsTable")
        self.assertEqual(ExcelProduct.Interface.excel_meta.key, "sku")
        self.assertEqual(set(ExcelProduct.Interface.excel_fields), {"sku", "name"})
        self.assertEqual(set(ExcelProduct.Interface.input_fields), {"sku"})

    def test_table_and_header_row_are_mutually_exclusive(self) -> None:
        with self.assertRaises(ValueError):

            class InvalidExcelProduct(GeneralManager):
                class Interface(ExcelInterface):
                    sku = ExcelCharField(unique=True)

                    class Meta:
                        workbook = "products.xlsx"
                        sheet = "Products"
                        table = "ProductsTable"
                        header_row = 1
                        key = "sku"


def write_product_workbook(
    path: Path,
    rows: list[list[object]],
    *,
    headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(headers or ["sku", "name"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def read_product_workbook(path: Path) -> list[dict[str, object | None]]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Products"]
        headers = [
            str(sheet.cell(row=1, column=column).value)
            for column in range(1, sheet.max_column + 1)
        ]
        records: list[dict[str, object | None]] = []
        for row_number in range(2, sheet.max_row + 1):
            record = {
                header: sheet.cell(row=row_number, column=column).value
                for column, header in enumerate(headers, start=1)
            }
            if all(value in (None, "") for value in record.values()):
                continue
            records.append(record)
        return records
    finally:
        workbook.close()


def product_workbook_value(
    path: Path,
    key: object,
    *,
    header: str = "name",
    key_header: str = "sku",
) -> object | None:
    for row in read_product_workbook(path):
        if row[key_header] == key:
            return row[header]
    return None


def product_workbook_key_count(
    path: Path,
    key: object,
    *,
    key_header: str = "sku",
) -> int:
    return sum(1 for row in read_product_workbook(path) if row[key_header] == key)


def set_product_workbook_value(
    path: Path,
    key: object,
    value: object,
    *,
    header: str = "name",
    key_header: str = "sku",
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Products"]
        headers = {
            str(sheet.cell(row=1, column=column).value): column
            for column in range(1, sheet.max_column + 1)
        }
        key_column = headers[key_header]
        value_column = headers[header]
        for row_number in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_number, column=key_column).value == key:
                sheet.cell(row=row_number, column=value_column).value = value
                workbook.save(path)
                return
        message = _missing_workbook_key_message(key)
        raise AssertionError(message)
    finally:
        workbook.close()


def append_product_workbook_row(path: Path, row: list[object]) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Products"]
        sheet.append(row)
        workbook.save(path)
    finally:
        workbook.close()


def delete_product_workbook_row(
    path: Path,
    key: object,
    *,
    key_header: str = "sku",
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Products"]
        headers = {
            str(sheet.cell(row=1, column=column).value): column
            for column in range(1, sheet.max_column + 1)
        }
        key_column = headers[key_header]
        for row_number in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_number, column=key_column).value == key:
                sheet.delete_rows(row_number, 1)
                workbook.save(path)
                return
        message = _missing_workbook_key_message(key)
        raise AssertionError(message)
    finally:
        workbook.close()


def _missing_workbook_key_message(key: object) -> str:
    return f"Workbook key {key!r} not found."


def build_product_manager(
    path: Path,
    *,
    sku_header: str | None = None,
    name_header: str | None = None,
) -> type[GeneralManager]:
    sku_field = ExcelCharField(header=sku_header, unique=True)
    name_field = ExcelCharField(header=name_header)

    class Product(GeneralManager):
        class Interface(ExcelInterface):
            sku = sku_field
            name = name_field

            class Meta:
                workbook = str(path)
                sheet = "Products"
                header_row = 1
                key = "sku"

    return Product


def build_parsed_key_product_manager(path: Path) -> type[GeneralManager]:
    def parse_sku(value: object) -> int:
        return int(str(value).removeprefix("SKU-"))

    def dump_sku(value: int | None) -> str | None:
        if value is None:
            return None
        return f"SKU-{value}"

    class Product(GeneralManager):
        class Interface(ExcelInterface):
            sku = ExcelField(
                int,
                parser=parse_sku,
                dumper=dump_sku,
                unique=True,
            )
            name = ExcelCharField()

            class Meta:
                workbook = str(path)
                sheet = "Products"
                header_row = 1
                key = "sku"

    return Product


class TempPathMixin:
    def setUp(self) -> None:
        super().setUp()  # type: ignore[misc]
        self._tempdir = TemporaryDirectory()
        self.addCleanup(self._tempdir.cleanup)  # type: ignore[attr-defined]

    def temp_path(self, name: str) -> Path:
        return Path(self._tempdir.name) / name


class ExcelStartupHookTests(TempPathMixin, SimpleTestCase):
    def setUp(self) -> None:
        super().setUp()
        clear_startup_hooks()
        clear_system_checks()

    def tearDown(self) -> None:
        clear_startup_hooks()
        clear_system_checks()
        super().tearDown()

    def test_startup_hook_registration_warms_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)

        Product.Interface.get_capabilities()
        hooks = registered_startup_hooks()
        self.assertIn(Product.Interface, hooks)
        self.assertEqual(len(hooks[Product.Interface]), 1)

        hooks[Product.Interface][0]()

        mirror = DEFAULT_EXCEL_STORE.mirror_for(Product.Interface)
        self.assertIn("SKU-1", mirror.rows)
        self.assertEqual(Product(sku="SKU-1").name, "Alpha")

    def test_startup_hook_logs_structure_error_without_raising(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()

        write_product_workbook(path, [], headers=["sku"])
        hooks = registered_startup_hooks()

        with self.assertLogs("general_manager.interface.excel", level="WARNING"):
            hooks[Product.Interface][0]()

        self.assertEqual(Product(sku="SKU-1").name, "Alpha")


class ExcelSystemCheckTests(TempPathMixin, SimpleTestCase):
    def setUp(self) -> None:
        super().setUp()
        clear_startup_hooks()
        clear_system_checks()

    def tearDown(self) -> None:
        clear_startup_hooks()
        clear_system_checks()
        super().tearDown()

    def _check_messages(self, manager_cls: type[GeneralManager]) -> list[object]:
        manager_cls.Interface.get_capabilities()
        checks = registered_system_checks()
        self.assertIn(manager_cls.Interface, checks)
        return [
            message for check in checks[manager_cls.Interface] for message in check()
        ]

    def test_system_check_returns_empty_list_for_valid_workbook(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)

        self.assertEqual(self._check_messages(Product), [])

    def test_system_check_reports_missing_declared_column_for_header_only_workbook(
        self,
    ) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [], headers=["sku"])
        Product = build_product_manager(path)

        messages = self._check_messages(Product)

        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0].id, "general_manager.excel.W001")
        self.assertIn("name", messages[0].msg)

    def test_system_check_reports_duplicate_and_blank_keys(self) -> None:
        cases = (
            ("duplicate.xlsx", [["SKU-1", "Alpha"], ["SKU-1", "Beta"]], "Duplicate"),
            ("blank.xlsx", [[None, "No SKU"]], "blank"),
        )
        for filename, rows, expected in cases:
            with self.subTest(filename=filename):
                clear_startup_hooks()
                clear_system_checks()
                path = self.temp_path(filename)
                write_product_workbook(path, rows)
                Product = build_product_manager(path)

                messages = self._check_messages(Product)

                self.assertEqual(len(messages), 1)
                self.assertEqual(messages[0].id, "general_manager.excel.W001")
                self.assertIn(expected, messages[0].msg)

    def test_system_check_reports_missing_custom_header(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [], headers=["sku", "name"])
        Product = build_product_manager(path, name_header="Product Name")

        messages = self._check_messages(Product)

        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0].id, "general_manager.excel.W001")
        self.assertIn("Product Name", messages[0].msg)

    def test_django_checks_include_multiple_nested_excel_interfaces(self) -> None:
        original_registered_checks = set(django_checks_registry.registered_checks)
        original_registered_interfaces = set(
            gm_bootstrap._registered_system_check_interfaces
        )
        self.addCleanup(
            setattr,
            django_checks_registry,
            "registered_checks",
            original_registered_checks,
        )
        self.addCleanup(
            setattr,
            gm_bootstrap,
            "_registered_system_check_interfaces",
            original_registered_interfaces,
        )
        django_checks_registry.registered_checks = set()
        gm_bootstrap._registered_system_check_interfaces.clear()

        first_path = self.temp_path("first.xlsx")
        second_path = self.temp_path("second.xlsx")
        write_product_workbook(first_path, [], headers=["sku"])
        write_product_workbook(second_path, [], headers=["sku"])
        FirstProduct = build_product_manager(first_path)
        SecondProduct = build_product_manager(second_path)
        # Distinct declared interfaces have distinct qualified names on current main.
        SecondProduct.Interface.__qualname__ += "Second"

        FirstProduct.Interface.get_capabilities()
        gm_bootstrap.register_system_checks()
        SecondProduct.Interface.get_capabilities()
        gm_bootstrap.register_system_checks()

        messages = [
            message
            for message in run_checks(tags=["general_manager"])
            if message.id == EXCEL_SYSTEM_CHECK_ID
        ]

        self.assertEqual(len(messages), 2)
        self.assertEqual(
            {message.obj for message in messages},
            {FirstProduct.Interface, SecondProduct.Interface},
        )


class ExcelSyncTests(TempPathMixin, SimpleTestCase):
    def test_sync_loads_rows_into_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(unique=True)
                name = ExcelCharField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        Product.sync_excel()

        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")

    def test_sync_uses_key_field_header_alias(self) -> None:
        path = self.temp_path("products.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["SKU", "name"])
        sheet.append(["SKU-1", "Alpha"])
        workbook.save(path)

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(header="SKU", unique=True)
                name = ExcelCharField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        Product.sync_excel()

        self.assertEqual(Product(sku="SKU-1").name, "Alpha")

    def test_missing_declared_column_keeps_last_valid_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(unique=True)
                name = ExcelCharField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        Product.sync_excel()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["sku"])
        sheet.append(["SKU-1"])
        workbook.save(path)

        with self.assertRaises(ExcelStructureError):
            Product.sync_excel()

        self.assertEqual(Product(sku="SKU-1").name, "Alpha")

    def test_header_only_missing_declared_column_keeps_last_valid_mirror(
        self,
    ) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()

        write_product_workbook(path, [], headers=["sku"])

        with self.assertRaises(ExcelStructureError) as error:
            Product.sync_excel()

        mirror = DEFAULT_EXCEL_STORE.mirror_for(Product.Interface)
        self.assertIs(mirror.structure_error, error.exception)
        self.assertIn("name", str(error.exception))
        self.assertEqual(Product(sku="SKU-1").name, "Alpha")

    def test_blank_raw_key_with_default_keeps_last_valid_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(required=False, default="DEFAULT-SKU", unique=True)
                name = ExcelCharField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        Product.sync_excel()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["sku", "name"])
        sheet.append([None, "Beta"])
        workbook.save(path)

        with self.assertRaises(ExcelStructureError) as error:
            Product.sync_excel()

        mirror = DEFAULT_EXCEL_STORE.mirror_for(Product.Interface)
        self.assertIs(mirror.structure_error, error.exception)
        self.assertEqual(Product(sku="SKU-1").name, "Alpha")

    def test_read_surfaces_validation_errors_without_stale_fallback(self) -> None:
        path = self.temp_path("products.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["sku", "qty"])
        sheet.append(["SKU-1", 5])
        workbook.save(path)

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(unique=True)
                qty = ExcelIntegerField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        Product.sync_excel()
        self.assertEqual(Product(sku="SKU-1").qty, 5)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["sku", "qty"])
        sheet.append(["SKU-1", "not-int"])
        workbook.save(path)

        with self.assertRaises(ExcelValidationError):
            _ = Product(sku="SKU-1").qty

    def test_sync_returns_delta_for_changes_and_noop(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])

        class Product(GeneralManager):
            class Interface(ExcelInterface):
                sku = ExcelCharField(unique=True)
                name = ExcelCharField()

                class Meta:
                    workbook = str(path)
                    sheet = "Products"
                    header_row = 1
                    key = "sku"

        created_delta = Product.Interface.sync_from_excel(force=True)
        self.assertIsInstance(created_delta, ExcelSyncDelta)
        self.assertEqual([row.key for row in created_delta.created], ["SKU-1"])
        self.assertEqual(created_delta.updated, ())
        self.assertEqual(created_delta.deleted, ())

        noop_delta = Product.Interface.sync_from_excel()
        self.assertIsInstance(noop_delta, ExcelSyncDelta)
        self.assertEqual(noop_delta.created, ())
        self.assertEqual(noop_delta.updated, ())
        self.assertEqual(noop_delta.deleted, ())

        write_product_workbook(path, [["SKU-1", "Alpha Prime"], ["SKU-2", "Beta"]])

        changed_delta = Product.Interface.sync_from_excel()
        self.assertEqual([row.key for row in changed_delta.created], ["SKU-2"])
        self.assertEqual(
            [
                (old.key, old.values["name"], new.values["name"])
                for old, new in changed_delta.updated
            ],
            [("SKU-1", "Alpha", "Alpha Prime")],
        )
        self.assertEqual(changed_delta.deleted, ())

        write_product_workbook(path, [["SKU-2", "Beta"]])

        deleted_delta = Product.Interface.sync_from_excel()
        self.assertEqual(deleted_delta.created, ())
        self.assertEqual(deleted_delta.updated, ())
        self.assertEqual([row.key for row in deleted_delta.deleted], ["SKU-1"])


class ExcelDependencyCacheInvalidationTests(TempPathMixin, SimpleTestCase):
    def setUp(self) -> None:
        super().setUp()
        cache.clear()
        self.addCleanup(cache.clear)

    def test_sync_update_invalidates_matching_filter_dependency_cache(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}

        @cached(cache="dependency")
        def alpha_count() -> int:
            calls["count"] += 1
            return Product.filter(name="Alpha").count()

        self.assertEqual(alpha_count(), 1)
        self.assertEqual(alpha_count(), 1)
        self.assertEqual(calls["count"], 1)

        set_product_workbook_value(path, "SKU-1", "Beta")
        Product.sync_excel()

        self.assertEqual(alpha_count(), 0)
        self.assertEqual(calls["count"], 2)

    def test_sync_update_enqueues_invalidated_cache_key_for_rewarm(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}

        @cached(cache="dependency")
        def alpha_count() -> int:
            calls["count"] += 1
            return Product.filter(name="Alpha").count()

        expected_key = make_cache_key(alpha_count, (), {})
        self.assertEqual(alpha_count(), 1)
        self.assertEqual(alpha_count(), 1)
        self.assertEqual(calls["count"], 1)

        enqueued_keys: list[tuple[str, ...]] = []

        def enqueue_rewarm(cache_keys: tuple[str, ...]) -> bool:
            self.assertFalse(is_dependency_data_change_active())
            enqueued_keys.append(tuple(cache_keys))
            return True

        set_product_workbook_value(path, "SKU-1", "Beta")
        with patch(
            "general_manager.api.graphql_warmup.enqueue_graphql_recipe_warmup",
            side_effect=enqueue_rewarm,
        ):
            Product.sync_excel()

        self.assertEqual(len(enqueued_keys), 1)
        self.assertIn(expected_key, enqueued_keys[0])

    def test_sync_create_invalidates_all_dependency_cache(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}

        @cached(cache="dependency")
        def product_count() -> int:
            calls["count"] += 1
            return Product.all().count()

        self.assertEqual(product_count(), 1)
        self.assertEqual(product_count(), 1)
        self.assertEqual(calls["count"], 1)

        append_product_workbook_row(path, ["SKU-2", "Beta"])
        Product.sync_excel()

        self.assertEqual(product_count(), 2)
        self.assertEqual(calls["count"], 2)

    def test_sync_delete_invalidates_all_dependency_cache(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Beta"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}

        @cached(cache="dependency")
        def product_count() -> int:
            calls["count"] += 1
            return Product.all().count()

        self.assertEqual(product_count(), 2)
        self.assertEqual(product_count(), 2)
        self.assertEqual(calls["count"], 1)

        delete_product_workbook_row(path, "SKU-2")
        Product.sync_excel()

        self.assertEqual(product_count(), 1)
        self.assertEqual(calls["count"], 2)

    def test_noop_sync_preserves_dependency_cache(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        generation = get_dependency_generation()
        calls = {"count": 0}

        @cached(cache="dependency")
        def alpha_count() -> int:
            calls["count"] += 1
            return Product.filter(name="Alpha").count()

        self.assertEqual(alpha_count(), 1)
        self.assertEqual(alpha_count(), 1)
        self.assertEqual(calls["count"], 1)

        Product.sync_excel()

        self.assertEqual(get_dependency_generation(), generation)
        self.assertEqual(alpha_count(), 1)
        self.assertEqual(calls["count"], 1)

    def test_nested_sync_bumps_generation_and_preserves_outer_barrier(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()

        begin_dependency_data_change()
        try:
            generation = get_dependency_generation()
            set_product_workbook_value(path, "SKU-1", "Beta")

            Product.sync_excel()

            self.assertEqual(get_dependency_generation(), generation + 1)
            self.assertTrue(is_dependency_data_change_active())
        finally:
            while is_dependency_data_change_active():
                end_dependency_data_change()

    def test_sync_during_dependency_compute_prevents_stale_publish(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}
        trigger_sync = {"needed": True}

        @cached(cache="dependency")
        def alpha_count() -> int:
            calls["count"] += 1
            result = Product.filter(name="Alpha").count()
            if trigger_sync["needed"]:
                trigger_sync["needed"] = False
                set_product_workbook_value(path, "SKU-1", "Beta")
                Product.sync_excel()
            return result

        expected_key = make_cache_key(alpha_count, (), {})

        self.assertEqual(alpha_count(), 1)
        self.assertIsNone(cache.get(expected_key))
        self.assertEqual(calls["count"], 1)

        self.assertEqual(alpha_count(), 0)
        self.assertEqual(alpha_count(), 0)
        self.assertEqual(calls["count"], 2)

    def test_nested_sync_during_dependency_compute_prevents_stale_publish(
        self,
    ) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        calls = {"count": 0}
        trigger_sync = {"needed": True}

        begin_dependency_data_change()
        try:

            @cached(cache="dependency")
            def alpha_count() -> int:
                calls["count"] += 1
                result = Product.filter(name="Alpha").count()
                if trigger_sync["needed"]:
                    trigger_sync["needed"] = False
                    set_product_workbook_value(path, "SKU-1", "Beta")
                    Product.sync_excel()
                    self.assertTrue(is_dependency_data_change_active())
                    end_dependency_data_change()
                return result

            expected_key = make_cache_key(alpha_count, (), {})

            self.assertEqual(alpha_count(), 1)
            self.assertFalse(is_dependency_data_change_active())
            self.assertIsNone(cache.get(expected_key))
            self.assertEqual(calls["count"], 1)

            self.assertEqual(alpha_count(), 0)
            self.assertEqual(alpha_count(), 0)
            self.assertEqual(calls["count"], 2)
        finally:
            while is_dependency_data_change_active():
                end_dependency_data_change()

    def test_noop_sync_does_not_enqueue_rewarm(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()

        @cached(cache="dependency")
        def alpha_count() -> int:
            return Product.filter(name="Alpha").count()

        self.assertEqual(alpha_count(), 1)

        with patch(
            "general_manager.api.graphql_warmup.enqueue_graphql_recipe_warmup",
            return_value=True,
        ) as enqueue_rewarm:
            Product.sync_excel()

        enqueue_rewarm.assert_not_called()

    def test_rewarm_enqueue_failure_is_logged_without_failing_sync(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Gamma"]])
        Product = build_product_manager(path)
        Product.sync_excel()

        @cached(cache="dependency")
        def alpha_count() -> int:
            return Product.filter(name="Alpha").count()

        self.assertEqual(alpha_count(), 1)
        set_product_workbook_value(path, "SKU-1", "Beta")

        with (
            patch(
                "general_manager.api.graphql_warmup.enqueue_graphql_recipe_warmup",
                side_effect=RuntimeError("boom"),
            ) as enqueue_rewarm,
            patch(
                "general_manager.interface.capabilities.excel.logger",
                create=True,
            ) as logger,
        ):
            Product.sync_excel()

        enqueue_rewarm.assert_called_once()
        logger.exception.assert_called_once_with("GraphQL warm-up requeue failed.")


class ExcelMutationTests(TempPathMixin, SimpleTestCase):
    def test_create_appends_row_and_updates_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)

        product = Product.create(sku="SKU-2", name="Beta", ignore_permission=True)

        self.assertEqual(product.identification, {"sku": "SKU-2"})
        self.assertEqual(product_workbook_value(path, "SKU-2"), "Beta")
        self.assertEqual(Product(sku="SKU-2").name, "Beta")

    def test_update_writes_row_when_excel_unchanged(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")

        result = product.update(name="Alpha Prime", ignore_permission=True)

        self.assertIs(result, product)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Alpha Prime")
        self.assertEqual(product.name, "Alpha Prime")

    def test_update_conflict_excel_wins_and_refreshes_mirror(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        set_product_workbook_value(path, "SKU-1", "Excel Edit")

        with self.assertRaisesRegex(ExcelWriteConflictError, "could not be saved"):
            product.update(name="GM Edit", ignore_permission=True)

        self.assertEqual(product_workbook_value(path, "SKU-1"), "Excel Edit")
        self.assertEqual(Product(sku="SKU-1").name, "Excel Edit")
        self.assertEqual(product.name, "Excel Edit")

    def test_stale_manager_update_conflict_after_intervening_sync(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")
        set_product_workbook_value(path, "SKU-1", "Excel Edit")
        Product.sync_excel()

        with self.assertRaisesRegex(ExcelWriteConflictError, "could not be saved"):
            product.update(name="GM Edit", ignore_permission=True)

        self.assertEqual(product_workbook_value(path, "SKU-1"), "Excel Edit")
        self.assertEqual(Product(sku="SKU-1").name, "Excel Edit")
        self.assertEqual(product.name, "Excel Edit")

    def test_update_rejects_key_field_change(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Beta"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")

        with self.assertRaisesRegex(
            ExcelMutationPayloadError,
            "Excel key field 'sku' cannot be updated",
        ):
            product.update(sku="SKU-2", ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 1)
        self.assertEqual(product_workbook_key_count(path, "SKU-2"), 1)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Alpha")
        self.assertEqual(Product(sku="SKU-1").name, "Alpha")
        self.assertEqual(product.name, "Alpha")

    def test_create_rejects_unknown_payload_field(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)

        with self.assertRaisesRegex(
            ExcelMutationPayloadError,
            "Unknown Excel field 'nmae'",
        ):
            Product.create(
                sku="SKU-2",
                name="Beta",
                nmae="Typo",
                ignore_permission=True,
            )

        self.assertEqual(product_workbook_key_count(path, "SKU-2"), 0)

    def test_update_rejects_unknown_payload_field(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")

        with self.assertRaisesRegex(
            ExcelMutationPayloadError,
            "Unknown Excel field 'nmae'",
        ):
            product.update(nmae="GM Edit", ignore_permission=True)

        self.assertEqual(product_workbook_value(path, "SKU-1"), "Alpha")
        self.assertEqual(Product(sku="SKU-1").name, "Alpha")
        self.assertEqual(product.name, "Alpha")

    def test_delete_removes_row_when_excel_unchanged(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"], ["SKU-2", "Beta"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")

        product.delete(ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 0)
        self.assertEqual(product_workbook_key_count(path, "SKU-2"), 1)

    def test_delete_conflict_excel_wins(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        set_product_workbook_value(path, "SKU-1", "Excel Edit")

        with self.assertRaisesRegex(ExcelWriteConflictError, "could not be saved"):
            product.delete(ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 1)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Excel Edit")
        self.assertEqual(Product(sku="SKU-1").name, "Excel Edit")

    def test_stale_manager_delete_conflict_after_intervening_sync(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)
        Product.sync_excel()
        product = Product(sku="SKU-1")
        self.assertEqual(product.name, "Alpha")
        set_product_workbook_value(path, "SKU-1", "Excel Edit")
        Product.sync_excel()

        with self.assertRaisesRegex(ExcelWriteConflictError, "could not be saved"):
            product.delete(ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 1)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Excel Edit")
        self.assertEqual(Product(sku="SKU-1").name, "Excel Edit")
        self.assertEqual(product.name, "Excel Edit")

    def test_create_conflict_existing_excel_key_wins(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [["SKU-1", "Alpha"]])
        Product = build_product_manager(path)

        with self.assertRaisesRegex(ExcelWriteConflictError, "already exists"):
            Product.create(sku="SKU-1", name="GM Edit", ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 1)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Alpha")

    def test_create_rejects_missing_declared_write_header_before_append(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [], headers=["sku"])
        Product = build_product_manager(path)

        with self.assertRaisesRegex(
            ExcelStructureError,
            "Missing Excel column for field 'name'",
        ):
            Product.create(sku="SKU-1", name="Alpha", ignore_permission=True)

        self.assertEqual(read_product_workbook(path), [])
        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 0)

    def test_create_rejects_missing_custom_write_header_before_append(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [], headers=["sku"])
        Product = build_product_manager(path, name_header="Product Name")

        with self.assertRaisesRegex(
            ExcelStructureError,
            "Missing Excel column for field 'name'",
        ):
            Product.create(sku="SKU-1", name="Alpha", ignore_permission=True)

        self.assertEqual(read_product_workbook(path), [])
        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 0)

    def test_create_uses_custom_field_headers(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(
            path,
            [["SKU-1", "Alpha"]],
            headers=["SKU", "Product Name"],
        )
        Product = build_product_manager(
            path,
            sku_header="SKU",
            name_header="Product Name",
        )

        product = Product.create(sku="SKU-2", name="Beta", ignore_permission=True)

        self.assertEqual(product.identification, {"sku": "SKU-2"})
        self.assertEqual(
            product_workbook_value(
                path,
                "SKU-2",
                key_header="SKU",
                header="Product Name",
            ),
            "Beta",
        )
        self.assertEqual(Product(sku="SKU-2").name, "Beta")

    def test_update_uses_dumped_custom_key_for_workbook_lookup(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [])
        Product = build_parsed_key_product_manager(path)
        Product.create(sku=1, name="Alpha", ignore_permission=True)

        Product(sku=1).update(name="Alpha Prime", ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 1)
        self.assertEqual(product_workbook_value(path, "SKU-1"), "Alpha Prime")
        self.assertEqual(Product(sku=1).name, "Alpha Prime")

    def test_delete_uses_dumped_custom_key_for_workbook_lookup(self) -> None:
        path = self.temp_path("products.xlsx")
        write_product_workbook(path, [])
        Product = build_parsed_key_product_manager(path)
        Product.create(sku=1, name="Alpha", ignore_permission=True)

        Product(sku=1).delete(ignore_permission=True)

        self.assertEqual(product_workbook_key_count(path, "SKU-1"), 0)
        self.assertNotIn(1, DEFAULT_EXCEL_STORE.mirror_for(Product.Interface).rows)
        self.assertEqual(Product.all().count(), 0)
