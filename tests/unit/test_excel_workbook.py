from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from unittest.mock import Mock, patch

from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.table import (  # type: ignore[import-untyped]
    Table,
    TableStyleInfo,
)

from general_manager.interface.excel import ExcelMeta, ExcelStructureError
from general_manager.interface.excel_workbook import ExcelWorkbookAdapter


def save_table_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["sku", "name", "price", "notes"])
    sheet.append(["SKU-1", "Alpha", 10, "preserve"])
    sheet.append(["SKU-2", "Beta", 20, "preserve"])
    table = Table(displayName="ProductsTable", ref="A1:D3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    sheet.add_table(table)
    workbook.save(path)


def save_plain_workbook(
    path: Path,
    rows: list[list[object | None]],
    *,
    sheet_name: str = "Products",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def table_refs(path: Path) -> tuple[str, str | None]:
    workbook = load_workbook(path)
    try:
        table = workbook["Products"].tables["ProductsTable"]
        auto_filter_ref = None if table.autoFilter is None else table.autoFilter.ref
        return table.ref, auto_filter_ref
    finally:
        workbook.close()


class FakeCell:
    def __init__(self, sheet: FakeSheet, row: int, column: int) -> None:
        self.sheet = sheet
        self.row = row
        self.column = column

    @property
    def value(self) -> object | None:
        return self.sheet.values.get((self.row, self.column))

    @value.setter
    def value(self, value: object | None) -> None:
        self.sheet.values[(self.row, self.column)] = value


class FakeSheet:
    def __init__(
        self,
        rows: list[list[object | None]],
        *,
        tables: dict[str, Any] | None = None,
    ) -> None:
        self.tables = tables or {}
        self.max_row = len(rows)
        self.max_column = max((len(row) for row in rows), default=0)
        self.values: dict[tuple[int, int], object | None] = {}
        for row_number, row in enumerate(rows, start=1):
            for column, value in enumerate(row, start=1):
                self.values[(row_number, column)] = value

    def cell(self, row: int, column: int) -> FakeCell:
        return FakeCell(self, row, column)

    def delete_rows(self, row_number: int, amount: int) -> None:
        values: dict[tuple[int, int], object | None] = {}
        for (row, column), value in self.values.items():
            if row < row_number:
                values[(row, column)] = value
            elif row >= row_number + amount:
                values[(row - amount, column)] = value
        self.values = values
        self.max_row = max(self.max_row - amount, 0)


class FakeWorkbook:
    def __init__(
        self,
        sheets: dict[str, FakeSheet],
        *,
        save_error: Exception | None = None,
    ) -> None:
        self._sheets = sheets
        self.sheetnames = list(sheets)
        self.close = Mock()
        self.save = Mock(side_effect=save_error)

    def __getitem__(self, sheet_name: str) -> FakeSheet:
        return self._sheets[sheet_name]


class TempPathMixin:
    def setUp(self) -> None:
        super().setUp()  # type: ignore[misc]
        self._tempdir = TemporaryDirectory()
        self.addCleanup(self._tempdir.cleanup)  # type: ignore[attr-defined]

    def temp_path(self, name: str) -> Path:
        return Path(self._tempdir.name) / name


class ExcelWorkbookAdapterTests(TempPathMixin, SimpleTestCase):
    def test_reads_table_rows_by_header(self) -> None:
        path = self.temp_path("products.xlsx")
        save_table_workbook(path)
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        rows = ExcelWorkbookAdapter(meta).read_rows()

        self.assertEqual(rows.headers["sku"], 1)
        self.assertEqual(rows.records[0].values["sku"], "SKU-1")
        self.assertEqual(rows.records[0].values["name"], "Alpha")

    def test_reads_header_row_mode(self) -> None:
        path = self.temp_path("plain.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["ignore"])
        sheet.append(["sku", "name"])
        sheet.append(["SKU-1", "Alpha"])
        workbook.save(path)
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=2, key="sku")

        rows = ExcelWorkbookAdapter(meta).read_rows()

        self.assertEqual(rows.records[0].row_number, 3)
        self.assertEqual(rows.records[0].values, {"sku": "SKU-1", "name": "Alpha"})

    def test_update_preserves_unknown_columns(self) -> None:
        path = self.temp_path("products.xlsx")
        save_table_workbook(path)
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        ExcelWorkbookAdapter(meta).update_row("SKU-1", {"name": "Changed"})

        workbook = load_workbook(path)
        try:
            sheet = workbook["Products"]
            self.assertEqual(sheet["B2"].value, "Changed")
            self.assertEqual(sheet["D2"].value, "preserve")
        finally:
            workbook.close()

    def test_append_expands_table_and_auto_filter_ranges(self) -> None:
        path = self.temp_path("products.xlsx")
        save_table_workbook(path)
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        ExcelWorkbookAdapter(meta).append_row(
            {"sku": "SKU-3", "name": "Gamma", "price": 30}
        )

        self.assertEqual(table_refs(path), ("A1:D4", "A1:D4"))

    def test_delete_shrinks_table_and_auto_filter_ranges(self) -> None:
        path = self.temp_path("products.xlsx")
        save_table_workbook(path)
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        ExcelWorkbookAdapter(meta).delete_row("SKU-2")

        self.assertEqual(table_refs(path), ("A1:D2", "A1:D2"))

    def test_append_after_delete_uses_next_table_row_without_gap(self) -> None:
        path = self.temp_path("products.xlsx")
        save_table_workbook(path)
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        adapter = ExcelWorkbookAdapter(meta)
        adapter.delete_row("SKU-1")
        adapter.append_row({"sku": "SKU-3", "name": "Gamma", "price": 30})

        workbook = load_workbook(path)
        try:
            sheet = workbook["Products"]
            table = sheet.tables["ProductsTable"]
            auto_filter_ref = None if table.autoFilter is None else table.autoFilter.ref
            self.assertEqual((table.ref, auto_filter_ref), ("A1:D3", "A1:D3"))
            self.assertEqual(sheet["A2"].value, "SKU-2")
            self.assertEqual(sheet["A3"].value, "SKU-3")
            self.assertIsNone(sheet["A4"].value)
        finally:
            workbook.close()

    def test_read_rejects_blank_header_row(self) -> None:
        path = self.temp_path("plain.xlsx")
        save_plain_workbook(path, [["ignore"], [None, ""], ["SKU-1", "Alpha"]])
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=2, key="sku")

        with self.assertRaises(ExcelStructureError):
            ExcelWorkbookAdapter(meta).read_rows()

    def test_read_rejects_out_of_range_header_row(self) -> None:
        path = self.temp_path("plain.xlsx")
        save_plain_workbook(path, [["sku", "name"]])
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=3, key="sku")

        with self.assertRaises(ExcelStructureError):
            ExcelWorkbookAdapter(meta).read_rows()

    def test_read_allows_missing_declared_key_header(self) -> None:
        path = self.temp_path("plain.xlsx")
        save_plain_workbook(path, [["id", "name"], ["SKU-1", "Alpha"]])
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=1, key="sku")

        rows = ExcelWorkbookAdapter(meta).read_rows()

        self.assertEqual(rows.headers, {"id": 1, "name": 2})
        self.assertEqual(rows.records[0].values, {"id": "SKU-1", "name": "Alpha"})

    def test_write_methods_raise_structure_error_for_missing_sheet(self) -> None:
        path = self.temp_path("plain.xlsx")
        save_plain_workbook(
            path, [["sku", "name"], ["SKU-1", "Alpha"]], sheet_name="Other"
        )
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=1, key="sku")

        for method_name, args in (
            ("update_row", ("SKU-1", {"name": "Changed"})),
            ("append_row", ({"sku": "SKU-2", "name": "Beta"},)),
            ("delete_row", ("SKU-1",)),
        ):
            with self.subTest(method_name=method_name):
                adapter = ExcelWorkbookAdapter(meta)
                method = getattr(adapter, method_name)
                with self.assertRaises(ExcelStructureError):
                    method(*args)

    def test_read_closes_workbook_when_sheet_is_missing(self) -> None:
        path = self.temp_path("fake.xlsx")
        path.write_bytes(b"fake workbook")
        workbook = FakeWorkbook({})
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        with patch(
            "general_manager.interface.excel_workbook.load_workbook",
            return_value=workbook,
        ):
            with self.assertRaises(ExcelStructureError):
                ExcelWorkbookAdapter(meta).read_rows()

        workbook.close.assert_called_once_with()

    def test_read_closes_workbook_when_table_is_missing(self) -> None:
        path = self.temp_path("fake.xlsx")
        path.write_bytes(b"fake workbook")
        workbook = FakeWorkbook({"Products": FakeSheet([["sku"], ["SKU-1"]])})
        meta = ExcelMeta(
            workbook=str(path),
            sheet="Products",
            table="ProductsTable",
            key="sku",
        )

        with patch(
            "general_manager.interface.excel_workbook.load_workbook",
            return_value=workbook,
        ):
            with self.assertRaises(ExcelStructureError):
                ExcelWorkbookAdapter(meta).read_rows()

        workbook.close.assert_called_once_with()

    def test_read_closes_workbook_when_declared_key_header_is_missing(self) -> None:
        path = self.temp_path("fake.xlsx")
        path.write_bytes(b"fake workbook")
        workbook = FakeWorkbook({"Products": FakeSheet([["name"], ["Alpha"]])})
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=1, key="sku")

        with patch(
            "general_manager.interface.excel_workbook.load_workbook",
            return_value=workbook,
        ):
            rows = ExcelWorkbookAdapter(meta).read_rows()

        self.assertEqual(rows.records[0].values, {"name": "Alpha"})
        workbook.close.assert_called_once_with()

    def test_write_methods_close_workbook_when_save_fails(self) -> None:
        path = self.temp_path("fake.xlsx")
        path.write_bytes(b"fake workbook")
        meta = ExcelMeta(workbook=str(path), sheet="Products", header_row=1, key="sku")

        for method_name, args in (
            ("update_row", ("SKU-1", {"name": "Changed"})),
            ("append_row", ({"sku": "SKU-2", "name": "Beta"},)),
            ("delete_row", ("SKU-1",)),
        ):
            with self.subTest(method_name=method_name):
                workbook = FakeWorkbook(
                    {"Products": FakeSheet([["sku", "name"], ["SKU-1", "Alpha"]])},
                    save_error=OSError("cannot save"),
                )
                with patch(
                    "general_manager.interface.excel_workbook.load_workbook",
                    return_value=workbook,
                ):
                    with self.assertRaises(OSError):
                        method = getattr(ExcelWorkbookAdapter(meta), method_name)
                        method(*args)

                workbook.close.assert_called_once_with()


def test_table_delete_preserves_cells_outside_table(tmp_path):
    path = tmp_path / "adjacent.xlsx"
    save_table_workbook(path)
    wb = load_workbook(path)
    wb["Products"]["F2"] = "Keep this note"
    wb.save(path)
    wb.close()
    adapter = ExcelWorkbookAdapter(
        ExcelMeta(
            workbook=str(path), sheet="Products", table="ProductsTable", key="sku"
        )
    )
    adapter.delete_row("SKU-1")
    wb = load_workbook(path)
    assert wb["Products"]["F2"].value == "Keep this note"
    assert wb["Products"]["A2"].value == "SKU-2"
    wb.close()


def test_failed_save_leaves_original_workbook_intact(tmp_path):
    path = tmp_path / "atomic.xlsx"
    save_table_workbook(path)
    original = path.read_bytes()
    adapter = ExcelWorkbookAdapter(
        ExcelMeta(
            workbook=str(path), sheet="Products", table="ProductsTable", key="sku"
        )
    )

    def broken_save(self, target):
        Path(target).write_bytes(b"partial file")
        message = "disk full"
        raise OSError(message)

    import pytest

    with (
        patch("openpyxl.workbook.workbook.Workbook.save", broken_save),
        pytest.raises(OSError),
    ):
        adapter.update_row("SKU-1", {"name": "Changed"})
    assert path.read_bytes() == original


def test_append_refuses_to_overwrite_content_below_table(tmp_path):
    import pytest

    path = tmp_path / "below.xlsx"
    save_table_workbook(path)
    wb = load_workbook(path)
    wb["Products"]["A4"] = "Unrelated summary"
    wb.save(path)
    wb.close()
    original = path.read_bytes()
    adapter = ExcelWorkbookAdapter(
        ExcelMeta(
            workbook=str(path), sheet="Products", table="ProductsTable", key="sku"
        )
    )
    with pytest.raises(ExcelStructureError):
        adapter.append_row({"sku": "SKU-3", "name": "Gamma"})
    assert path.read_bytes() == original


def test_atomic_save_preserves_workbook_symlink(tmp_path):
    target = tmp_path / "target.xlsx"
    link = tmp_path / "link.xlsx"
    save_table_workbook(target)
    link.symlink_to(target)
    adapter = ExcelWorkbookAdapter(
        ExcelMeta(
            workbook=str(link), sheet="Products", table="ProductsTable", key="sku"
        )
    )
    adapter.update_row("SKU-1", {"name": "Changed"})
    assert link.is_symlink()
    wb = load_workbook(target)
    assert wb["Products"]["B2"].value == "Changed"
    wb.close()
