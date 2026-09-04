"""Workbook access for ExcelInterface."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from functools import wraps
import os
from tempfile import NamedTemporaryFile
from hashlib import sha256
from pathlib import Path
from typing import Any, Callable, TypeVar, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.formula.tokenizer import Token  # type: ignore[import-untyped]
from openpyxl.formula.translate import (  # type: ignore[import-untyped]
    Translator,
    TranslatorError,
)
from openpyxl.utils.cell import (  # type: ignore[import-untyped]
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.filters import AutoFilter  # type: ignore[import-untyped]

from general_manager.interface.excel import (
    ExcelMeta,
    ExcelStructureError,
    ExcelWriteConflictError,
)


class _ExcelWorkbookStructureError(ExcelStructureError):
    @classmethod
    def occupied_expansion(cls) -> _ExcelWorkbookStructureError:
        return cls("Cannot expand Excel table over existing worksheet content.")

    @classmethod
    def missing_sheet(cls, sheet: str) -> _ExcelWorkbookStructureError:
        return cls(f"Workbook sheet {sheet!r} is missing.")

    @classmethod
    def missing_table(cls, table: str) -> _ExcelWorkbookStructureError:
        return cls(f"Excel table {table!r} is missing.")

    @classmethod
    def missing_header_row(cls) -> _ExcelWorkbookStructureError:
        return cls("Excel header_row is not configured.")

    @classmethod
    def invalid_header_row(cls, row: int) -> _ExcelWorkbookStructureError:
        return cls(f"Excel header row {row} is blank or outside the workbook range.")

    @classmethod
    def missing_key_column(cls, key: str) -> _ExcelWorkbookStructureError:
        return cls(f"Key column {key!r} is missing.")


@dataclass(frozen=True, slots=True)
class WorkbookFingerprint:
    mtime_ns: int
    size: int
    digest: str


@dataclass(frozen=True, slots=True)
class ExcelWorkbookRecord:
    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class ExcelWorkbookRows:
    headers: dict[str, int]
    records: tuple[ExcelWorkbookRecord, ...]
    fingerprint: WorkbookFingerprint


def workbook_fingerprint(path: str) -> WorkbookFingerprint:
    file_path = Path(path)
    stat = file_path.stat()
    digest = sha256(file_path.read_bytes()).hexdigest()
    return WorkbookFingerprint(stat.st_mtime_ns, stat.st_size, digest)


def _translate_formula(formula: str, *, origin: str, target: str) -> str:
    """Translate a formula, replacing references moved out of bounds with #REF!."""
    translator = Translator(formula, origin=origin)
    try:
        return cast(str, translator.translate_formula(target))
    except TranslatorError:
        origin_row, origin_column = coordinate_to_tuple(origin)
        target_row, target_column = coordinate_to_tuple(target)
        row_delta = target_row - origin_row
        column_delta = target_column - origin_column
        translated = ["="]
        for token in translator.get_tokens():
            if token.type == Token.OPERAND and token.subtype == Token.RANGE:
                try:
                    translated.append(
                        Translator.translate_range(
                            token.value,
                            row_delta,
                            column_delta,
                        )
                    )
                except TranslatorError:
                    translated.append("#REF!")
            else:
                translated.append(token.value)
        return "".join(translated)


_F = TypeVar("_F", bound=Callable[..., Any])


def _workbook_locked(method: _F) -> _F:
    @wraps(method)
    def locked(self: ExcelWorkbookAdapter, *args: Any, **kwargs: Any) -> Any:
        from general_manager.interface.excel_store import DEFAULT_EXCEL_STORE

        with DEFAULT_EXCEL_STORE.lock_for(self.meta.workbook):
            return method(self, *args, **kwargs)

    return cast(_F, locked)


class ExcelWorkbookAdapter:
    def __init__(self, meta: ExcelMeta) -> None:
        self.meta = meta

    @_workbook_locked
    def read_rows(self) -> ExcelWorkbookRows:
        fingerprint = workbook_fingerprint(self.meta.workbook)
        workbook = load_workbook(self.meta.workbook, data_only=True)
        try:
            sheet = self._sheet(workbook)
            min_col, min_row, max_col, max_row = self._range_bounds(sheet)
            headers = self._read_headers(sheet, min_row, min_col, max_col)
            records: list[ExcelWorkbookRecord] = []
            for row_number in range(min_row + 1, max_row + 1):
                values = {
                    header: sheet.cell(row=row_number, column=column).value
                    for header, column in headers.items()
                }
                if all(value in (None, "") for value in values.values()):
                    continue
                records.append(
                    ExcelWorkbookRecord(row_number=row_number, values=values)
                )
            if workbook_fingerprint(self.meta.workbook) != fingerprint:
                raise ExcelWriteConflictError.workbook_changed("read")
            return ExcelWorkbookRows(
                headers=headers,
                records=tuple(records),
                fingerprint=fingerprint,
            )
        finally:
            workbook.close()

    @_workbook_locked
    def update_row(
        self,
        key: object,
        values: dict[str, Any],
        *,
        expected_fingerprint: WorkbookFingerprint | None = None,
    ) -> None:
        expected = workbook_fingerprint(self.meta.workbook)
        if expected_fingerprint is not None and expected != expected_fingerprint:
            raise ExcelWriteConflictError.workbook_changed("update")
        workbook = load_workbook(self.meta.workbook)
        try:
            sheet = self._sheet(workbook)
            min_col, min_row, max_col, max_row = self._range_bounds(sheet)
            headers = self._read_validated_headers(sheet, min_row, min_col, max_col)
            row_number = self._find_row_number(sheet, headers, min_row, max_row, key)
            for header, value in values.items():
                column = headers.get(header)
                if column is not None:
                    sheet.cell(row=row_number, column=column).value = value
            self._save(workbook, expected)
        finally:
            workbook.close()

    @_workbook_locked
    def append_row(self, values: dict[str, Any]) -> None:
        expected = workbook_fingerprint(self.meta.workbook)
        workbook = load_workbook(self.meta.workbook)
        try:
            sheet = self._sheet(workbook)
            min_col, min_row, max_col, max_row = self._range_bounds(sheet)
            headers = self._read_validated_headers(sheet, min_row, min_col, max_col)
            row_number = max_row + 1
            if self.meta.table and any(
                sheet.cell(row=row_number, column=column).value is not None
                for column in range(min_col, max_col + 1)
            ):
                raise _ExcelWorkbookStructureError.occupied_expansion()
            for header, column in headers.items():
                sheet.cell(row=row_number, column=column).value = values.get(header)
            self._resize_table_if_needed(sheet, min_col, min_row, max_col, row_number)
            self._save(workbook, expected)
        finally:
            workbook.close()

    @_workbook_locked
    def delete_row(
        self,
        key: object,
        *,
        expected_fingerprint: WorkbookFingerprint | None = None,
    ) -> None:
        expected = workbook_fingerprint(self.meta.workbook)
        if expected_fingerprint is not None and expected != expected_fingerprint:
            raise ExcelWriteConflictError.workbook_changed("delete")
        workbook = load_workbook(self.meta.workbook)
        try:
            sheet = self._sheet(workbook)
            min_col, min_row, max_col, max_row = self._range_bounds(sheet)
            headers = self._read_validated_headers(sheet, min_row, min_col, max_col)
            row_number = self._find_row_number(sheet, headers, min_row, max_row, key)
            # Shift only the managed rectangle, preserving adjacent worksheet data.
            for row in range(row_number, max_row):
                for column in range(min_col, max_col + 1):
                    source = sheet.cell(row=row + 1, column=column)
                    target = sheet.cell(row=row, column=column)
                    value = source.value
                    if isinstance(value, str) and value.startswith("="):
                        value = _translate_formula(
                            value,
                            origin=source.coordinate,
                            target=target.coordinate,
                        )
                    target.value = value
                    target._style = copy(source._style)
                    target.comment = copy(source.comment)
                    target.hyperlink = copy(source.hyperlink)
            for column in range(min_col, max_col + 1):
                sheet.cell(row=max_row, column=column).value = None
                sheet.cell(row=max_row, column=column).comment = None
                sheet.cell(row=max_row, column=column).hyperlink = None
            self._resize_table_if_needed(sheet, min_col, min_row, max_col, max_row - 1)
            self._save(workbook, expected)
        finally:
            workbook.close()

    def _save(self, workbook: Any, expected: WorkbookFingerprint) -> None:
        path = Path(self.meta.workbook).resolve()
        # A failed ZIP save cannot truncate the authoritative workbook.
        with NamedTemporaryFile(
            dir=path.parent, prefix=f".{path.name}.", suffix=".xlsx", delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
        try:
            workbook.save(temporary_path)
            with temporary_path.open("rb") as stream:
                os.fsync(stream.fileno())
            if workbook_fingerprint(str(path)) != expected:
                raise ExcelWriteConflictError.workbook_changed("save")
            temporary_path.chmod(path.stat().st_mode)
            os.replace(temporary_path, path)
        finally:
            temporary_path.unlink(missing_ok=True)

    def _sheet(self, workbook: Any) -> Any:
        if self.meta.sheet not in workbook.sheetnames:
            raise _ExcelWorkbookStructureError.missing_sheet(self.meta.sheet)
        return workbook[self.meta.sheet]

    def _range_bounds(self, sheet: Any) -> tuple[int, int, int, int]:
        if self.meta.table:
            table = sheet.tables.get(self.meta.table)
            if table is None:
                raise _ExcelWorkbookStructureError.missing_table(self.meta.table)
            return cast(tuple[int, int, int, int], range_boundaries(table.ref))
        header_row = self.meta.header_row
        if header_row is None:
            raise _ExcelWorkbookStructureError.missing_header_row()
        max_row = sheet.max_row
        while max_row > header_row:
            if any(
                sheet.cell(row=max_row, column=column).value not in (None, "")
                for column in range(1, sheet.max_column + 1)
            ):
                break
            max_row -= 1
        return 1, header_row, sheet.max_column, max_row

    @staticmethod
    def _read_headers(
        sheet: Any, row: int, min_col: int, max_col: int
    ) -> dict[str, int]:
        if row > sheet.max_row:
            raise _ExcelWorkbookStructureError.invalid_header_row(row)
        headers: dict[str, int] = {}
        for column in range(min_col, max_col + 1):
            value = sheet.cell(row=row, column=column).value
            if value not in (None, ""):
                headers[str(value)] = column
        if not headers:
            raise _ExcelWorkbookStructureError.invalid_header_row(row)
        return headers

    def _read_validated_headers(
        self, sheet: Any, row: int, min_col: int, max_col: int
    ) -> dict[str, int]:
        headers = self._read_headers(sheet, row, min_col, max_col)
        if self.meta.key not in headers:
            raise _ExcelWorkbookStructureError.missing_key_column(self.meta.key)
        return headers

    def _find_row_number(
        self,
        sheet: Any,
        headers: dict[str, int],
        min_row: int,
        max_row: int,
        key: object,
    ) -> int:
        key_column = headers.get(self.meta.key)
        if key_column is None:
            raise _ExcelWorkbookStructureError.missing_key_column(self.meta.key)
        for row_number in range(min_row + 1, max_row + 1):
            if sheet.cell(row=row_number, column=key_column).value == key:
                return row_number
        raise KeyError(key)

    def _resize_table_if_needed(
        self,
        sheet: Any,
        min_col: int,
        min_row: int,
        max_col: int,
        max_row: int,
    ) -> None:
        if not self.meta.table:
            return
        table = sheet.tables[self.meta.table]
        table_ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        table.ref = table_ref
        if table.autoFilter is None:
            table.autoFilter = AutoFilter(ref=table_ref)
        else:
            table.autoFilter.ref = table_ref
