from __future__ import annotations
from datetime import datetime
from typing import Any, ClassVar
import os

from general_manager.interface.baseInterface import (
    InterfaceBase,
    classPostCreationMethod,
    classPreCreationMethod,
    generalManagerClassName,
    attributes,
    interfaceBaseClass,
    newlyCreatedGeneralManagerClass,
    newlyCreatedInterfaceClass,
    relatedClass,
    AttributeTypedDict,
)
from general_manager.manager.input import Input
from general_manager.interface.excelDataField import ExcelDataField


class ExcelInterface(InterfaceBase):
    """Interface for importing and exporting data via Excel files."""

    _interface_type = "excel"
    _file_path: ClassVar[str | None] = None
    _file_mtime: ClassVar[float | None] = None
    _rows: ClassVar[list[dict[str, Any]]] = []
    _data_cache: ClassVar[dict[tuple, dict[str, Any]]] = {}
    _sheet_name: ClassVar[str | None] = None
    input_fields: dict[str, Input]
    data_fields: dict[str, ExcelDataField]

    def getData(self, search_date: datetime | None = None) -> dict[str, Any]:
        """Return data for the row matching this interface's identification."""

        cls = self.__class__

        if cls._file_path is None:
            raise ValueError("No Excel file configured for this interface.")

        current_mtime = os.path.getmtime(cls._file_path)
        if cls._file_mtime != current_mtime:
            cls._file_mtime = current_mtime
            cls._data_cache.clear()
            cls._rows = cls._load_rows()

        if not cls._rows:
            cls._rows = cls._load_rows()

        cache_key = tuple(self.identification.get(name) for name in cls.input_fields.keys())
        if cache_key in cls._data_cache:
            return cls._data_cache[cache_key]

        for row in cls._rows:
            row_key = tuple(row.get(name) for name in cls.input_fields.keys())
            if row_key == cache_key:
                cls._data_cache[cache_key] = row
                return row

        raise KeyError("Matching row not found in Excel file.")

    @classmethod
    def _load_rows(cls) -> list[dict[str, Any]]:
        """Load all rows from the configured Excel file."""

        from openpyxl import load_workbook

        workbook = load_workbook(cls._file_path, data_only=True)
        if cls._sheet_name is not None:
            sheet = workbook[cls._sheet_name]
        else:
            sheet = workbook.active

        header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        col_index = {name: header.index(name) for name in cls.data_fields.keys() if name in header}

        rows: list[dict[str, Any]] = []
        for row_cells in sheet.iter_rows(min_row=2, values_only=True):
            row_data: dict[str, Any] = {}
            for name, field in cls.data_fields.items():
                idx = col_index.get(name)
                value = row_cells[idx] if idx is not None else None
                if value is None:
                    value = field.default
                row_data[name] = field.cast(value) if value is not None else None
            rows.append(row_data)
        return rows

    @classmethod
    def getAttributeTypes(cls) -> dict[str, AttributeTypedDict]:  # pragma: no cover - abstract
        raise NotImplementedError

    @classmethod
    def getAttributes(cls) -> dict[str, Any]:  # pragma: no cover - abstract
        raise NotImplementedError

    @classmethod
    def filter(cls, **kwargs: Any):  # pragma: no cover - abstract
        raise NotImplementedError

    @classmethod
    def exclude(cls, **kwargs: Any):  # pragma: no cover - abstract
        raise NotImplementedError

    @classmethod
    def getFieldType(cls, field_name: str) -> type:  # pragma: no cover - abstract
        raise NotImplementedError

    @staticmethod
    def _preCreate(
        name: generalManagerClassName,
        attrs: attributes,
        interface: interfaceBaseClass,
    ) -> tuple[attributes, interfaceBaseClass, None]:
        input_fields: dict[str, Input[Any]] = {}
        data_fields: dict[str, ExcelDataField] = {}
        file_path: str | None = None
        sheet_name: str | None = None

        for key, value in vars(interface).items():
            if key.startswith("__"):
                continue
            if isinstance(value, Input):
                input_fields[key] = value
            elif isinstance(value, ExcelDataField):
                data_fields[key] = value
            elif key == "Meta" and isinstance(value, type):
                if hasattr(value, "file_path"):
                    file_path = getattr(value, "file_path")
                if hasattr(value, "sheet_name"):
                    sheet_name = getattr(value, "sheet_name")

        attrs["_interface_type"] = interface._interface_type
        interface_attrs = {
            "input_fields": input_fields,
            "data_fields": data_fields,
            "_file_mtime": None,
            "_rows": [],
            "_data_cache": {},
            "_sheet_name": sheet_name,
        }
        if file_path is not None:
            interface_attrs["_file_path"] = file_path
        interface_cls = type(interface.__name__, (interface,), interface_attrs)
        attrs["Interface"] = interface_cls
        return attrs, interface_cls, None

    @staticmethod
    def _postCreate(
        new_class: newlyCreatedGeneralManagerClass,
        interface_class: newlyCreatedInterfaceClass,
        model: relatedClass,
    ) -> None:
        interface_class._parent_class = new_class

    @classmethod
    def handleInterface(cls) -> tuple[classPreCreationMethod, classPostCreationMethod]:
        """Return handler functions used by the meta class during creation."""
        return cls._preCreate, cls._postCreate
